import click
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


@click.command()
@click.option('--capitalize',default='',help="will capitalise the contents of the worksheet")
def cli(capitalize):
    click.echo("hey there")
    book = load_workbook("C:/PythonCourse/txt.xlsx")
    list = book.get_sheet_names()
    for s in list:
        sheet = book.get_sheet_by_name(s)
        for rowno in range(1, sheet.max_row + 1):
            for colno in range(1, sheet.max_column + 1):
                name = sheet.cell(row=rowno, column=colno).value
                sheet.cell(row=rowno, column=colno).value = name.upper()
        book.save('meena1.xlsx')
        click.echo("Done")






