from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

book = load_workbook("C:/PythonCourse/txt.xlsx")
op=load_workbook("C:/PythonCourse/output.xlsx")
ops=op.active
list = book.get_sheet_names()
list.remove("Sheet3")
for s in list:
    sheet = book.get_sheet_by_name(s)
    for rowno in range(1, sheet.max_row):
        for colno in range(1, sheet.max_column + 1):
            name = sheet.cell(row=rowno, column=colno).value
            sheet.cell(row=rowno, column=colno).value = name.upper()
    book.save('meena1.xlsx')