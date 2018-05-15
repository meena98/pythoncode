import click
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

@click.command()
@click.argument('name')
#@click.option('--copy',default='',help="copies data from html to a excel sheet")
def cli(**kwargs):
    output='{0}'.format(kwargs['name'])
    click.echo("hey there")
    wb = Workbook()
    ws = wb.create_sheet()
    page = requests.get("https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html")
    soup = BeautifulSoup(page.content, 'html.parser')
    l = []
    count = 0
    for temp in soup.select('td'):
        if count == 7:
            count = 0
            del l[0]
            ws.append(l)
            l = []
        l.append(temp.text)
        count = count + 1
    wb.save(output)
    click.echo("done")