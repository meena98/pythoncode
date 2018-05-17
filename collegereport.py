import smtplib
import click
import _mysql
import xlrd
import  MySQLdb
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


wb = Workbook()
ws = wb.create_sheet()
dbPassword = "Mysql!@#"
database = MySQLdb.connect (host="localhost", user = "meena", passwd = dbPassword, db = "student")
query = "SELECT sinfo.dbnames,sinfo.college,marks.marks1,marks.marks2,marks.marks3,marks.maximum,marks.total FROM sinfo INNER JOIN marks ON sinfo.dbnames=marks.dbnames WHERE sinfo.college='%s'" % ("anits")
cursor = database.cursor()
cursor.execute(query)
result1=cursor.fetchall()
print(result1)
l=["NAME","COLLEGE","MARKS1","MARKS2","MARKS3","MARKS4","TOTAL"]
ws.append(l)
l=[]
count=0
for temp1 in result1:
    for temp in temp1:
        if count==7:
            count=0
            ws.append(l)
            l=[]
        l.append(temp)
        count+=1
query = "SELECT sinfo.college,count(sinfo.college),sum(marks1),sum(marks2),sum(marks3),sum(maximum),sum(total) FROM sinfo INNER JOIN marks ON sinfo.dbnames=marks.dbnames WHERE sinfo.college='%s'" % ("anits")
cursor.execute(query)
result = cursor.fetchone()
minimum = min(result[2], result[3], result[4],result[5])
maximum = max(result[2], result[3], result[4],result[5])
total = result[6]/4
ws.append([""])
ws.append(["COLLEGE SUMMARY"])
ws.append(["COLLEGE","COUNT","MIN","MAX","AVG"])
ws.append([result[0],result[1],minimum,maximum,total])
print(result[0], result[1], minimum, total, maximum)
query = "SELECT count(sinfo.college),min(total),max(total),sum(total) FROM sinfo INNER JOIN marks ON sinfo.dbnames=marks.dbnames"
cursor.execute(query)
result = cursor.fetchone()
total = result[3]/103
ws.append([""])
ws.append(["GLOBAL SUMMARY"])
ws.append(["COUNT","MIN","MAX","AVG"])
print(result[0],result[1], result[2], total)
ws.append([result[0],result[1], result[2], total])
wb.save("db.xlsx")
cursor.close()
database.commit()
database.close()
fromaddr = "s.meena9874@gmail.com"
toaddr = "aruna1976mmv@gmail.com"
msg = MIMEMultipart()
msg['From'] = fromaddr
msg['To'] = toaddr
msg['Subject'] = "Subject of the Mail"
body = "Body_of_the_mail"
msg.attach(MIMEText(body, 'plain'))
filename = "db.xlsx"
attachment = open("C:\PythonCourse\db.xlsx", "rb")
p = MIMEBase('application', 'octet-stream')
p.set_payload((attachment).read())
encoders.encode_base64(p)
p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
msg.attach(p)
s = smtplib.SMTP('smtp.gmail.com', 587)
s.starttls()
s.login(fromaddr, "")
text = msg.as_string()
s.sendmail(fromaddr, toaddr, text)
s.quit()

